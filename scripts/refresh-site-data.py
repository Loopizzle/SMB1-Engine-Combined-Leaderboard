import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import math
import os
import re
import tempfile
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SPREADSHEET_ID = "1gsIoLL-rWP5lhpiAmyCGOeZmruj1nuJYXHMp1nP95xQ"
REQUIRED_SHEETS = (
    "Combined Leaderboard",
    "Runs",
    "Boards",
    "Games",
    "Monthly View",
    "Monthly Winners",
    "Yearly View",
    "Yearly Winners",
    "About",
)
MONTHLY_CHAMPION_START = "2015-01"


def rows_from(sheet, start=1):
    rows = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index < start or not any(value is not None for value in row):
            continue
        rows.append(list(row))
    return rows


def dictionaries(rows):
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        for row in rows[1:]
    ]


def compact(item, field_map):
    return {target: item.get(source) for source, target in field_map.items()}


def sheet_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def round2(value):
    return math.floor(float(value) * 100 + 0.5) / 100


def placement_strength(place, board_size):
    if place <= 0 or board_size <= 0:
        return 0
    if board_size == 1:
        return 1 if place == 1 else 0
    return max(0, 1 - ((place - 1) / max(1, board_size - 1)))


def board_weight(board_size):
    return math.pow(math.log2(board_size + 1), 1.35) if board_size > 0 else 0


def monthly_performance_score(place, monthly_runs, current_board_size):
    strength = placement_strength(place, monthly_runs)
    if strength <= 0:
        return 0
    depth = max(current_board_size, monthly_runs)
    weight = max(1, board_weight(depth))
    competition = math.log2(monthly_runs + 1)
    score = math.pow(strength, 1.5) * weight * max(1, competition) * 2
    if place == 1:
        score += weight * 1.5
    elif place == 2:
        score += weight * 0.8
    elif place == 3:
        score += weight * 0.4
    return score


def monthly_board_key(board_key):
    parts = str(board_key or "").split("|")
    if len(parts) < 5:
        return ""
    return "|".join((parts[0], parts[1], parts[2], "|".join(parts[4:])))


def historical_board_index(boards):
    index = {}
    for board in boards:
        key = monthly_board_key(board.get("boardKey"))
        if not key:
            continue
        abbreviation = str(board.get("gameAbbr") or "")
        subcategory = str(board.get("subcategory") or "")
        toggle = f"{abbreviation} (Luigi)" if abbreviation in {"annsmb", "smbtll", "smbtllce"} and re.search(r"\bluigi\b", subcategory, re.IGNORECASE) else abbreviation
        index[key] = {
            "included": bool(board.get("included")),
            "boardKey": str(board.get("boardKey") or ""),
            "gameAbbr": abbreviation,
            "gameToggle": toggle,
            "game": str(board.get("game") or abbreviation),
            "scope": str(board.get("scope") or ""),
            "category": str(board.get("category") or ""),
            "level": board.get("level"),
            "subcategory": board.get("subcategory"),
            "runCount": int(board.get("runCount") or 0),
        }
    return index


def historical_monthly_history(sheet, boards, games):
    if sheet is None:
        return [], []

    enabled_games = {str(game.get("abbr")): bool(game.get("included")) for game in games}
    board_index = historical_board_index(boards)

    history = dictionaries(rows_from(sheet))
    runs = []
    for item in history:
        month = str(item.get("Verified At") or "")[:7]
        run_id = sheet_text(item.get("Run ID"))
        player_key = sheet_text(item.get("Player Key"))
        board_key = str(item.get("Board Key") or "")
        seconds = item.get("Seconds")
        if month < MONTHLY_CHAMPION_START or not re.fullmatch(r"\d{4}-\d{2}", month):
            continue
        if not run_id or not player_key or not board_key or seconds in (None, ""):
            continue
        board = board_index.get(board_key)
        if board and (not board["included"] or not enabled_games.get(board["gameToggle"], True)):
            continue
        runs.append({
            "month": month,
            "runId": run_id,
            "playerKey": player_key,
            "runner": sheet_text(item.get("Runner")) or player_key,
            "country": str(item.get("Country") or ""),
            "profile": str(item.get("Profile") or ""),
            "boardKey": board_key,
            "gameId": str(item.get("Game ID") or board_key.split("|", 1)[0]),
            "seconds": float(seconds),
        })

    board_runs = {}
    for run in runs:
        group = board_runs.setdefault((run["month"], run["boardKey"]), {})
        group.setdefault(run["runId"], run)

    performance = {}
    for (_, board_key), group in board_runs.items():
        ranked = sorted(group.values(), key=lambda run: (run["seconds"], run["runId"]))
        current_size = board_index.get(board_key, {}).get("runCount", len(ranked))
        for index, run in enumerate(ranked, start=1):
            performance[run["runId"]] = monthly_performance_score(index, len(ranked), current_size)

    stats = {}
    seen = set()
    for run in runs:
        dedupe_key = (run["runId"], run["playerKey"])
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        key = (run["month"], run["playerKey"])
        stat = stats.setdefault(key, {
            "runner": run["runner"],
            "country": run["country"],
            "profile": run["profile"],
            "performance": 0,
            "runs": 0,
            "boards": set(),
            "games": set(),
        })
        stat["performance"] += performance.get(run["runId"], 0)
        stat["runs"] += 1
        stat["boards"].add(run["boardKey"])
        stat["games"].add(run["gameId"])

    players_by_month = {}
    for (month, _), stat in stats.items():
        volume = math.sqrt(stat["runs"]) * 8
        variety = max(0, len(stat["games"]) - 1) * 10
        players_by_month.setdefault(month, []).append({
            "runner": stat["runner"],
            "country": stat["country"],
            "totalScore": round2(stat["performance"] + stat["runs"] + volume + variety),
            "performanceScore": round2(stat["performance"]),
            "runs": stat["runs"],
            "volumeBonus": round2(volume),
            "varietyBonus": round2(variety),
            "uniqueBoards": len(stat["boards"]),
            "uniqueGames": len(stat["games"]),
            "profile": stat["profile"],
        })

    rows = []
    winners = []
    for month in sorted(players_by_month):
        players = sorted(players_by_month[month], key=lambda player: (
            -player["totalScore"],
            -player["performanceScore"],
            -player["runs"],
            -player["uniqueBoards"],
            player["runner"],
        ))
        for rank, player in enumerate(players, start=1):
            rows.append({"month": month, "rank": rank, **player})
        champion = players[0]
        winners.append({
            "Month": month,
            "Winner": champion["runner"],
            "Flag": None,
            "Country": champion["country"],
            "Total Score": champion["totalScore"],
            "Performance Score": champion["performanceScore"],
            "Verified Runs": champion["runs"],
            "Volume Bonus": champion["volumeBonus"],
            "Variety Bonus": champion["varietyBonus"],
            "Unique Boards": champion["uniqueBoards"],
            "Unique Games": champion["uniqueGames"],
            "Profile": champion["profile"],
        })

    return rows, winners


def historical_career_runs(sheet, boards):
    if sheet is None:
        return []

    board_index = historical_board_index(boards)
    result = []
    seen = set()
    for item in dictionaries(rows_from(sheet)):
        run_id = sheet_text(item.get("Run ID")).strip()
        player_key = sheet_text(item.get("Player Key")).strip()
        runner = (sheet_text(item.get("Runner")) or player_key).strip()
        board = board_index.get(str(item.get("Board Key") or ""))
        verified_at = str(item.get("Verified At") or "").strip()
        run_date = str(item.get("Run Date") or "").strip()[:10] or verified_at[:10]
        seconds = item.get("Seconds")
        dedupe_key = (run_id, player_key)
        if not run_id or not player_key or not runner or not board or seconds in (None, ""):
            continue
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", run_date) or run_date < "2015-01-01":
            continue
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        result.append({
            "id": run_id,
            "playerKey": player_key,
            "runner": runner,
            "country": str(item.get("Country") or ""),
            "profile": str(item.get("Profile") or ""),
            "boardKey": board["boardKey"],
            "gameAbbr": board["gameAbbr"],
            "gameToggle": board["gameToggle"],
            "game": board["game"],
            "scope": board["scope"],
            "category": board["category"],
            "level": board["level"],
            "subcategory": board["subcategory"],
            "seconds": float(seconds),
            "runDate": run_date,
            "verifiedAt": verified_at or None,
            "runLink": str(item.get("Run Link") or ""),
        })
    return result


def flag_url(formula):
    if not isinstance(formula, str):
        return None
    match = re.search(r'https?://[^";,]+', formula)
    return match.group(0) if match else None


def download_workbook(spreadsheet_id, destination):
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    request = urllib.request.Request(url, headers={"User-Agent": "smb1ecl-daily-refresh/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        content = response.read()
    if len(content) < 100_000 or not content.startswith(b"PK"):
        raise RuntimeError("Google Sheets returned an invalid or unexpectedly small workbook")
    destination.write_bytes(content)


def local_flag_filename(url):
    if not isinstance(url, str):
        return None
    match = re.fullmatch(r"https://(?:www\.)?speedrun\.com/images/flags/([a-z0-9/_-]+)\.png", url, re.IGNORECASE)
    return f"{match.group(1).replace('/', '-')}.png" if match else None


def country_flag_urls():
    source = (ROOT / "app" / "page.tsx").read_text(encoding="utf-8")
    block = re.search(r"const countryCodes:[^{]+\{(.*?)\n\};", source, re.DOTALL)
    if not block:
        raise RuntimeError("Could not read the website country flag mapping")
    entries = re.findall(r"(?:'([^']+)'|([A-Za-z]+))\s*:\s*'([^']+)'", block.group(1))
    return {
        quoted or bare: f"https://www.speedrun.com/images/flags/{code}.png"
        for quoted, bare, code in entries
    }


def cache_flag_assets(payload):
    fallback_urls = country_flag_urls()
    for player in payload["combined"]:
        if not player.get("Flag URL"):
            player["Flag URL"] = fallback_urls.get(player.get("Country"))

    flag_targets = {
        player["Flag URL"]: local_flag_filename(player.get("Flag URL"))
        for player in payload["combined"]
        if local_flag_filename(player.get("Flag URL"))
    }
    if not flag_targets:
        print("No Speedrun.com flag assets were found in the workbook")
        return

    output_dir = ROOT / "public" / "flags"
    output_dir.mkdir(parents=True, exist_ok=True)

    def download(item):
        url, filename = item
        request = urllib.request.Request(url, headers={"User-Agent": "smb1ecl-daily-refresh/1.0"})
        with urllib.request.urlopen(request, timeout=30) as response:
            content = response.read()
        if len(content) < 50 or not content.startswith(b"\x89PNG\r\n\x1a\n"):
            raise RuntimeError(f"Invalid PNG response for {url}")
        destination = output_dir / filename
        temporary = destination.with_suffix(".png.tmp")
        temporary.write_bytes(content)
        temporary.replace(destination)
        return url, f"./flags/{filename}"

    local_urls = {}
    failures = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(download, item): item[0] for item in flag_targets.items()}
        for future in as_completed(futures):
            try:
                url, local_url = future.result()
                local_urls[url] = local_url
            except Exception as error:
                failures.append(f"{futures[future]}: {error}")

    for player in payload["combined"]:
        player["Flag URL"] = local_urls.get(player.get("Flag URL"), player.get("Flag URL"))

    print(f"Cached {len(local_urls)} unique flag assets; {len(failures)} remained remote")
    for failure in failures:
        print(f"Flag cache warning: {failure}")


def extract_payload(workbook_path):
    values = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    formulas = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        missing = [name for name in REQUIRED_SHEETS if name not in values.sheetnames]
        if missing:
            raise RuntimeError(f"Workbook is missing required sheets: {', '.join(missing)}")

        combined = dictionaries(rows_from(values["Combined Leaderboard"]))
        combined_flags = dictionaries(rows_from(formulas["Combined Leaderboard"]))
        for index, player in enumerate(combined):
            player["Runner"] = sheet_text(player.get("Runner"))
            if "Player Key" in player:
                player["Player Key"] = sheet_text(player.get("Player Key"))
            formula = combined_flags[index].get("Flag") if index < len(combined_flags) else None
            player["Flag URL"] = flag_url(formula)

        runs = [
            compact(item, {
                "Run ID": "id",
                "Included Now": "included",
                "Game Abbr": "gameAbbr",
                "Game": "game",
                "Scope": "scope",
                "Category": "category",
                "Level": "level",
                "Subcategory Values": "subcategory",
                "Place": "place",
                "Board Size": "boardSize",
                "Runner": "runner",
                "Country": "country",
                "Time": "time",
                "Seconds": "seconds",
                "Platform": "platform",
                "Console/Emulator": "hardware",
                "Run Date": "runDate",
                "Verified At": "verifiedAt",
                "Run Link": "runLink",
                "Player Key": "playerKey",
                "Profile": "profile",
                "Performance Points": "performancePoints",
                "Medal Points": "medalPoints",
                "Depth Points": "depthPoints",
                "WR Credit": "wrCredit",
                "Top 3 Credit": "top3Credit",
                "Top 10 Credit": "top10Credit",
                "Board Key": "boardKey",
                "Game Toggle": "gameToggle",
            })
            for item in dictionaries(rows_from(values["Runs"]))
        ]
        for run in runs:
            for field in ("id", "runner", "playerKey", "boardKey", "gameAbbr", "gameToggle"):
                run[field] = sheet_text(run.get(field))

        boards = [
            compact(item, {
                "Included": "included",
                "Game Abbr": "gameAbbr",
                "Game": "game",
                "Scope": "scope",
                "Category": "category",
                "Level": "level",
                "Subcategory Values": "subcategory",
                "Run Count": "runCount",
                "Timing": "timing",
                "Leaderboard Link": "leaderboardLink",
                "Board Key": "boardKey",
            })
            for item in dictionaries(rows_from(values["Boards"]))
        ]

        games = [
            compact(item, {
                "Included": "included",
                "Game Abbr": "abbr",
                "Game": "name",
                "Game ID": "id",
                "Leaderboard Link": "leaderboardLink",
            })
            for item in dictionaries(rows_from(values["Games"]))
        ]

        monthly = []
        for row in rows_from(values["Monthly View"]):
            if len(row) < 39 or row[26] in (None, "Month"):
                continue
            monthly.append({
                "month": row[26],
                "rank": row[27],
                "runner": row[28],
                "country": row[30],
                "totalScore": row[31],
                "performanceScore": row[32],
                "runs": row[33],
                "volumeBonus": row[34],
                "varietyBonus": row[35],
                "uniqueBoards": row[36],
                "uniqueGames": row[37],
                "profile": row[38],
            })

        yearly = []
        for row in rows_from(values["Yearly View"]):
            if len(row) < 39 or row[26] in (None, "Year"):
                continue
            yearly.append({
                "year": row[26],
                "rank": row[27],
                "runner": row[28],
                "country": row[30],
                "totalScore": row[31],
                "performanceScore": row[32],
                "runs": row[33],
                "volumeBonus": row[34],
                "varietyBonus": row[35],
                "uniqueBoards": row[36],
                "uniqueGames": row[37],
                "profile": row[38],
            })

        history_sheet = values["Yearly Data"] if "Yearly Data" in values.sheetnames else None
        archived_monthly, archived_monthly_winners = historical_monthly_history(history_sheet, boards, games)
        career_runs = historical_career_runs(history_sheet, boards) if history_sheet is not None else runs
        current_months = {str(row.get("month") or "")[:7] for row in monthly}
        monthly = [
            row for row in archived_monthly
            if str(row.get("month") or "")[:7] not in current_months
        ] + monthly
        monthly.sort(key=lambda row: (str(row.get("month") or "")[:7], int(row.get("rank") or 0)))

        current_monthly_winners = dictionaries(rows_from(values["Monthly Winners"]))
        monthly_winners_by_period = {
            str(row.get("Month") or "")[:7]: row
            for row in archived_monthly_winners
        }
        # The cache is complete for older years, while the visible sheet has the
        # authoritative current months gathered by the latest Apps Script refresh.
        for row in current_monthly_winners:
            monthly_winners_by_period[str(row.get("Month") or "")[:7]] = row
        monthly_winners = [
            monthly_winners_by_period[period]
            for period in sorted(monthly_winners_by_period)
            if period
        ]

        return {
            "generatedFrom": "The SMB1 Engine Combined Leaderboard public workbook",
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "combined": combined,
            "runs": runs,
            "careerRuns": career_runs,
            "boards": boards,
            "games": games,
            "monthly": monthly,
            "yearly": yearly,
            "monthlyWinners": monthly_winners,
            "yearlyWinners": dictionaries(rows_from(values["Yearly Winners"])),
            "about": dictionaries(rows_from(values["About"])),
        }
    finally:
        values.close()
        formulas.close()


def validate_payload(payload):
    minimums = {
        "combined": 100,
        "runs": 1_000,
        "careerRuns": 1_000,
        "boards": 20,
        "games": 6,
        "monthly": 1,
        "yearly": 1,
        "monthlyWinners": 1,
        "yearlyWinners": 1,
    }
    problems = [
        f"{name} has {len(payload[name])} rows; expected at least {minimum}"
        for name, minimum in minimums.items()
        if len(payload.get(name, [])) < minimum
    ]
    if any(not run.get("id") or not run.get("boardKey") for run in payload["runs"]):
        problems.append("one or more runs are missing a run ID or board key")
    if problems:
        raise RuntimeError("Refusing to publish invalid workbook data: " + "; ".join(problems))


def main():
    parser = argparse.ArgumentParser(description="Refresh website data from the public Google Sheet")
    parser.add_argument("--workbook", type=Path, help="Use a local workbook instead of downloading it")
    parser.add_argument("--output", type=Path, default=ROOT / "public" / "site-data.json")
    args = parser.parse_args()

    spreadsheet_id = os.environ.get("SMB1ECL_SPREADSHEET_ID", DEFAULT_SPREADSHEET_ID)
    with tempfile.TemporaryDirectory() as temp_dir:
        workbook_path = args.workbook or Path(temp_dir) / "smb1ecl.xlsx"
        if args.workbook is None:
            download_workbook(spreadsheet_id, workbook_path)
        payload = extract_payload(workbook_path)
        validate_payload(payload)
        cache_flag_assets(payload)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    temporary_output = args.output.with_suffix(args.output.suffix + ".tmp")
    temporary_output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str),
        encoding="utf-8",
    )
    temporary_output.replace(args.output)
    print(json.dumps({key: len(value) for key, value in payload.items() if isinstance(value, list)}, indent=2))


if __name__ == "__main__":
    main()
